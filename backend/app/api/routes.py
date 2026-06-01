"""
API 路由 - 扩展版
支持公司文件夹上传、数据抽取、文书生成
"""
import os
import uuid
import shutil
import zipfile
import asyncio
from typing import List, Optional
from datetime import datetime

from fastapi import APIRouter, UploadFile, File, HTTPException, Form, BackgroundTasks
from fastapi.responses import FileResponse
from pydantic import BaseModel

from ..services.pdf_service import PDFService
from ..services.excel_service import ExcelService
from ..services.docx_service import DocxService
from ..services.extractor_service import ExtractorService
from ..services.database_service import DatabaseService
from ..services.ai_service import stream_chat_response, chat_response, analyze_file_content, get_form_field_suggestion, check_form_completeness
from ..core.agent import LegalAgent
from ..models.case import CaseData, CompanyInfo, DefendantInfo, DebtInfo, CaseInfo
from ..models.loan import LoanContracts

# FastAPI响应
from fastapi.responses import StreamingResponse

router = APIRouter()

# 服务实例
pdf_service = PDFService()
excel_service = ExcelService()
docx_service = DocxService()
extractor_service = ExtractorService()
database_service = DatabaseService()
legal_agent = LegalAgent()

# 目录配置
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "../../uploads")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "../../outputs")
TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), "../../../诉讼文书模板")

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(os.path.join(OUTPUT_DIR, "cases"), exist_ok=True)


# ==================== 请求模型 ====================

class CaseInfo(BaseModel):
    """案件信息模型"""
    id: str
    filename: str
    plaintiff: str = ""
    defendant: str = ""
    case_type: str = ""
    amount: str = ""
    facts: str = ""
    evidence: List[str] = []
    court: str = ""
    status: str = "pending"
    created_at: str = ""
    raw_text: str = ""


class DocumentRequest(BaseModel):
    """文书生成请求"""
    case_ids: List[str]
    doc_type: str  # 起诉状, 证据目录, 申请书


class SelectedDocumentsRequest(BaseModel):
    """选中文书类型生成请求"""
    case_ids: List[str]
    doc_types: List[str]  # 选中的文书类型列表


class UpdateCaseRequest(BaseModel):
    """更新案件请求"""
    company_info: Optional[dict] = None
    defendants: Optional[List[dict]] = None
    debt_info: Optional[dict] = None
    case_info: Optional[dict] = None
    lawyer_letter_info: Optional[dict] = None  # 律师函信息
    lawer: Optional[str] = None


# ==================== 原有接口（保持兼容） ====================

@router.post("/upload")
async def upload_files(files: List[UploadFile] = File(...)):
    """上传 PDF 文件或 ZIP 压缩包（旧版接口）"""
    uploaded_cases = []

    for file in files:
        file_path = os.path.join(UPLOAD_DIR, file.filename)

        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        if file.filename.lower().endswith('.zip'):
            with zipfile.ZipFile(file_path, 'r') as zip_ref:
                for name in zip_ref.namelist():
                    if name.lower().endswith('.pdf'):
                        pdf_path = os.path.join(UPLOAD_DIR, os.path.basename(name))
                        with zip_ref.open(name) as source, open(pdf_path, 'wb') as target:
                            shutil.copyfileobj(source, target)

                        case = await process_pdf(pdf_path, os.path.basename(name))
                        uploaded_cases.append(case)
        else:
            case = await process_pdf(file_path, file.filename)
            uploaded_cases.append(case)

    return {"success": True, "cases": uploaded_cases}


async def process_pdf(file_path: str, filename: str) -> dict:
    """处理 PDF 文件并提取案件信息"""
    case_id = str(uuid.uuid4())[:8]
    raw_text = pdf_service.extract_text(file_path)
    case_data = await legal_agent.parse_case(raw_text)

    case_info = {
        "id": case_id,
        "filename": filename,
        "plaintiff": case_data.get("plaintiff", ""),
        "defendant": case_data.get("defendant", ""),
        "case_type": case_data.get("case_type", ""),
        "amount": case_data.get("amount", ""),
        "facts": case_data.get("facts", ""),
        "evidence": case_data.get("evidence", []),
        "court": case_data.get("court", ""),
        "status": "pending",
        "created_at": datetime.now().isoformat(),
        "raw_text": raw_text[:500]
    }

    return case_info


@router.get("/cases")
async def get_cases():
    """获取所有案件列表"""
    # 从JSON文件加载
    cases = database_service.list_cases()
    return {"cases": cases}


@router.get("/cases/{case_id}")
async def get_case(case_id: str):
    """获取单个案件详情"""
    case_data = database_service.load_case(case_id)
    if case_data is None:
        raise HTTPException(status_code=404, detail="案件不存在")
    return case_data.model_dump()


@router.put("/cases/{case_id}")
async def update_case(case_id: str, request: UpdateCaseRequest):
    """更新案件信息"""
    case_data = database_service.load_case(case_id)
    if case_data is None:
        raise HTTPException(status_code=404, detail="案件不存在")

    # 更新字段
    if request.company_info:
        for key, value in request.company_info.items():
            if hasattr(case_data.company_info, key):
                setattr(case_data.company_info, key, value)

    if request.defendants:
        case_data.defendants = [DefendantInfo(**d) for d in request.defendants]

    if request.debt_info:
        for key, value in request.debt_info.items():
            if hasattr(case_data.debt_info, key):
                setattr(case_data.debt_info, key, value)

    if request.case_info:
        for key, value in request.case_info.items():
            if hasattr(case_data.case_info, key):
                setattr(case_data.case_info, key, value)

    if request.lawyer_letter_info:
        for key, value in request.lawyer_letter_info.items():
            if hasattr(case_data.lawyer_letter_info, key):
                setattr(case_data.lawyer_letter_info, key, value)

    if request.lawer:
        case_data.lawer = request.lawer

    # 保存
    database_service.save_case(case_data)

    return {"success": True, "case": case_data.model_dump()}


@router.delete("/cases/{case_id}")
async def delete_case(case_id: str):
    """删除案件"""
    if not database_service.delete_case(case_id):
        raise HTTPException(status_code=404, detail="案件不存在")
    return {"success": True}


# ==================== 新增接口 ====================

@router.post("/ocr/idcard")
async def ocr_idcard(file: UploadFile = File(...)):
    """OCR识别身份证图片

    接收图片文件，返回识别的身份证信息
    """
    try:
        # 保存上传的图片
        temp_path = os.path.join(UPLOAD_DIR, f"temp_{file.filename}")
        with open(temp_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        # 使用OCR识别
        result = legal_agent.ocr_image_local(temp_path)

        # 清理临时文件
        os.remove(temp_path)

        return {"success": True, "data": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"OCR识别失败: {str(e)}")


@router.post("/ocr/idcard-pdf")
async def ocr_idcard_pdf(file: UploadFile = File(...)):
    """OCR识别身份证PDF

    接收PDF文件，转换为图片后进行OCR识别
    """
    try:
        # 保存上传的PDF
        temp_path = os.path.join(UPLOAD_DIR, f"temp_{file.filename}")
        with open(temp_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        # 将PDF转换为图片
        images = pdf_service.pdf_to_images(temp_path, output_dir=os.path.join(OUTPUT_DIR, "temp"))

        result = {"def_name": "", "def_gender": "", "def_nation": "", "def_id": "", "def_addr": ""}

        if images:
            # 对第一页进行OCR
            result = legal_agent.ocr_image_local(images[0])

            # 清理临时图片
            for img_path in images:
                try:
                    os.remove(img_path)
                except:
                    pass

        # 清理临时PDF
        os.remove(temp_path)

        return {"success": True, "data": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"OCR识别失败: {str(e)}")


@router.get("/company/folders")
async def get_company_folders():
    """获取已有的公司文件夹列表"""
    # 使用固定的公司文件汇总目录
    project_dir = os.path.join(os.path.dirname(__file__), "../../../公司文件汇总")
    project_dir = os.path.abspath(project_dir)

    folders = []

    # 扫描目录
    for item in os.listdir(project_dir):
        item_path = os.path.join(project_dir, item)
        if os.path.isdir(item_path):
            # 检查是否是公司文件夹（格式：ID_公司名 或直接公司名）
            if "_" in item:
                parts = item.split("_", 1)
                if len(parts) == 2:
                    folder_id = parts[0]
                    company_name = parts[1]
                else:
                    folder_id = item[:8]  # 使用前8个字符作为ID
                    company_name = item
            else:
                folder_id = item[:8]  # 使用前8个字符作为ID
                company_name = item

            # 使用extractor_service统计文件（会过滤干扰文件）
            stats = extractor_service.get_file_stats(item_path)

            folders.append({
                "id": folder_id,
                "name": company_name,
                "path": item_path,
                "file_count": stats["total_files"],
                "file_types": {k: len(v) for k, v in stats["by_type"].items()},
                "ignored_count": stats.get("ignored_count", 0)
            })

    return {"folders": folders}


# 文件分类关键词
FILE_PATTERNS = {
    "public_report": ["公示系统", "企业信用信息公示"],
    "id_card_front": ["身份证正面照片", "身份证正面"],
    "id_card_back": ["身份证反面照片", "身份证反面"],
    "quota_contract": ["额度合同"],
    "loan_contract": ["借款合同"],
    "loan_flow": ["放款流水", "放款存证"],
    "repay_flow": ["还款流水"],
    "guarantee_contract": ["担保合同"],
    "calc_logic": ["金额计算逻辑"]
}

# 干扰文件关键词
IGNORE_PATTERNS = [
    "个人信息使用授权书", "个人信息处理授权书", "个人征信授权书", "个人授权委托书",
    "企业信息使用授权书", "企业信息处理授权书", "企业征信授权书", "企业授权委托书", "企业授权协议书",
    "授信申请声明书", "仲裁申请书", "送达地址确认书", "送达地址线索书",
    "法定代表人身份证明书", "营业执照", "借款合同编号",
    "提前到期后还款计划表", "证据目录", "合同类映射表",
    ".zip", "保证合同", "原告送达"
]


def classify_file_type(filename: str) -> str:
    """根据文件名分类文件类型（排除干扰文件）"""
    # 检查是否需要忽略
    for pattern in IGNORE_PATTERNS:
        if pattern in filename:
            return "ignored"

    for file_type, keywords in FILE_PATTERNS.items():
        for keyword in keywords:
            if keyword in filename:
                return file_type
    return "other"


@router.post("/company/process-folder")
async def process_existing_folder(folder_path: str = Form(...)):
    """处理已有的公司文件夹"""
    if not os.path.exists(folder_path):
        raise HTTPException(status_code=404, detail="文件夹不存在")

    try:
        case_data = await process_company_folder(folder_path)
        database_service.save_case(case_data)

        return {
            "success": True,
            "case": case_data.model_dump(),
            "message": f"成功处理文件夹"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"处理失败: {str(e)}")


@router.post("/company/upload")
async def upload_company_folder(files: List[UploadFile] = File(...)):
    """上传公司文件夹（支持多文件）

    接收一个公司文件夹中的所有文件，自动分类并提取数据
    """
    # 创建临时文件夹
    temp_dir = os.path.join(UPLOAD_DIR, f"temp_{uuid.uuid4().hex[:8]}")
    os.makedirs(temp_dir, exist_ok=True)

    try:
        # 保存所有文件（处理文件夹上传时的路径问题）
        for file in files:
            # 获取文件名（可能包含相对路径）
            filename = file.filename
            # 只取文件名部分，忽略路径
            if "/" in filename or "\\" in filename:
                filename = os.path.basename(filename.replace("\\", "/"))

            # 跳过临时文件和目录
            if filename.startswith("~$") or filename.startswith("."):
                continue

            file_path = os.path.join(temp_dir, filename)

            # 确保目录存在
            os.makedirs(os.path.dirname(file_path) if os.path.dirname(file_path) else temp_dir, exist_ok=True)

            with open(file_path, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)

        # 处理文件夹
        case_data = await process_company_folder(temp_dir)

        # 保存到数据库
        database_service.save_case(case_data)

        return {
            "success": True,
            "case": case_data.model_dump(),
            "message": f"成功处理 {len(files)} 个文件"
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"处理失败: {str(e)}")
    finally:
        # 清理临时文件夹（可选）
        pass  # 保留文件供后续处理


async def process_company_folder(folder_path: str) -> CaseData:
    """处理公司文件夹，提取所有数据"""
    # 使用抽取服务处理（主要从Excel提取数据）
    case_data = extractor_service.process_company_folder(folder_path, legal_agent)

    # 解析公示系统PDF（提取住所、登记机关等）
    for filename in os.listdir(folder_path):
        if "公示系统" in filename or "企业信用信息公示" in filename:
            pdf_path = os.path.join(folder_path, filename)
            print(f"[公示系统] 正在解析: {filename}")

            # 先尝试直接提取文本
            pdf_text = pdf_service.extract_text(pdf_path)

            if not pdf_text.strip():
                # 如果文本为空（扫描版PDF），使用OCR
                print(f"[公示系统] 文本为空，使用OCR")
                pdf_text = pdf_service.extract_text_with_ocr(pdf_path)

            if pdf_text:
                # 使用规则解析（更快更准确）
                public_info = legal_agent._parse_public_report_rules(pdf_text)

                if public_info:
                    # 更新公司信息
                    if public_info.get("target_company"):
                        case_data.company_info.target_company = public_info["target_company"]
                    if public_info.get("company_capital"):
                        case_data.company_info.company_capital = public_info["company_capital"]
                    if public_info.get("company_establish"):
                        case_data.company_info.company_establish = public_info["company_establish"]
                    if public_info.get("company_addr"):
                        case_data.company_info.company_addr = public_info["company_addr"]
                    if public_info.get("company_reg"):
                        case_data.company_info.company_reg = public_info["company_reg"]
                    if public_info.get("legal_representative"):
                        case_data.company_info.legal_representative = public_info["legal_representative"]
                    if public_info.get("company_cancel_apply"):
                        case_data.company_info.company_cancel_apply = public_info["company_cancel_apply"]
                    if public_info.get("cancel_doc"):
                        case_data.company_info.cancel_doc = public_info["cancel_doc"]
                    if public_info.get("capital_status"):
                        case_data.company_info.capital_status = public_info["capital_status"]
                    if public_info.get("subscribe_date"):
                        case_data.company_info.subscribe_date = public_info["subscribe_date"]

                    # 更新股东信息
                    shareholders = public_info.get("shareholders", [])
                    for i, sh in enumerate(shareholders):
                        if i < len(case_data.defendants):
                            case_data.defendants[i].def_share = sh.get("share", "")
                            if not case_data.defendants[i].def_name:
                                case_data.defendants[i].def_name = sh.get("name", "")
                        else:
                            # 添加新股东
                            new_def = DefendantInfo(
                                def_name=sh.get("name", ""),
                                def_share=sh.get("share", "")
                            )
                            case_data.defendants.append(new_def)

                    print(f"[公示系统] 解析完成: {public_info.get('target_company', '')}")
            break  # 只处理第一个公示系统文件

    # 解析身份证PDF（提取性别、民族、住址等）
    for filename in os.listdir(folder_path):
        if "身份证正面" in filename:
            pdf_path = os.path.join(folder_path, filename)
            print(f"[身份证] 正在解析: {filename}")

            # 尝试提取文本
            pdf_text = pdf_service.extract_text(pdf_path)

            if not pdf_text.strip():
                pdf_text = pdf_service.extract_text_with_ocr(pdf_path)

            if pdf_text:
                id_info = legal_agent._parse_id_card_rules(pdf_text)

                if id_info:
                    # 从文件名提取姓名
                    name_from_file = ""
                    if "_身份证" in filename:
                        name_from_file = filename.split("_身份证")[0].split("_")[-1]

                    # 匹配被告
                    matched = False
                    for def_info in case_data.defendants:
                        if id_info.get("def_id") and def_info.def_id == id_info.get("def_id"):
                            def_info.def_gender = id_info.get("def_gender", "") or def_info.def_gender
                            def_info.def_nation = id_info.get("def_nation", "") or def_info.def_nation
                            def_info.def_addr = id_info.get("def_addr", "") or def_info.def_addr
                            matched = True
                            break
                        if (id_info.get("def_name") or name_from_file) and def_info.def_name == (id_info.get("def_name") or name_from_file):
                            def_info.def_gender = id_info.get("def_gender", "") or def_info.def_gender
                            def_info.def_nation = id_info.get("def_nation", "") or def_info.def_nation
                            def_info.def_id = id_info.get("def_id", "") or def_info.def_id
                            def_info.def_addr = id_info.get("def_addr", "") or def_info.def_addr
                            matched = True
                            break

                    if not matched and (id_info.get("def_name") or name_from_file):
                        case_data.defendants.append(DefendantInfo(
                            def_name=id_info.get("def_name") or name_from_file,
                            def_gender=id_info.get("def_gender", ""),
                            def_nation=id_info.get("def_nation", ""),
                            def_id=id_info.get("def_id", ""),
                            def_addr=id_info.get("def_addr", "")
                        ))

                    print(f"[身份证] 解析完成: {id_info.get('def_name', '')}")
            break  # 只处理第一个身份证文件

    # 推荐管辖法院
    if case_data.defendants and case_data.defendants[0].def_addr:
        case_data.case_info.court_name = legal_agent.recommend_court(case_data.defendants[0].def_addr)

    # 设置案由
    case_data.case_info.case_reason = "清算责任纠纷"

    return case_data


@router.get("/company/{case_id}/data")
async def get_company_data(case_id: str):
    """获取公司JSON数据"""
    case_data = database_service.load_case(case_id)
    if case_data is None:
        raise HTTPException(status_code=404, detail="案件不存在")

    return {
        "success": True,
        "data": case_data.model_dump()
    }


@router.post("/company/{case_id}/export")
async def export_company_excel(case_id: str):
    """导出公司要素Excel"""
    case_data = database_service.load_case(case_id)
    if case_data is None:
        raise HTTPException(status_code=404, detail="案件不存在")

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = os.path.join(OUTPUT_DIR, f"database_element_{case_id}_{timestamp}.xlsx")

    excel_service.generate_database_element(case_data, output_path)

    return FileResponse(
        output_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"database_element_{case_data.company_info.target_company or case_id}.xlsx"
    )


@router.post("/company/export-all")
async def export_all_cases():
    """导出所有案件汇总Excel"""
    cases = []
    for case_summary in database_service.list_cases():
        case_data = database_service.load_case(case_summary["id"])
        if case_data:
            cases.append(case_data)

    if not cases:
        raise HTTPException(status_code=400, detail="没有案件数据")

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = os.path.join(OUTPUT_DIR, f"案件汇总_{timestamp}.xlsx")

    excel_service.generate_batch_excel(cases, output_path)

    return FileResponse(
        output_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"案件汇总_{timestamp}.xlsx"
    )


@router.get("/templates")
async def get_templates():
    """获取可用模板列表"""
    templates = []

    # 自动扫描模板目录
    if os.path.exists(TEMPLATE_DIR):
        for filename in sorted(os.listdir(TEMPLATE_DIR)):
            if filename.endswith('.docx') and not filename.startswith('~$'):
                # 从文件名生成显示名称
                name = filename.replace('.docx', '').replace('模板', '')
                if name.startswith(('1-', '2-', '3-', '4-', '5-', '6-', '7-', '8-', '9-', '10-', '11-', '12-', '13-', '14-', '15-')):
                    name = name[2:]  # 去掉编号前缀
                templates.append({
                    "name": name,
                    "filename": filename,
                    "path": os.path.join(TEMPLATE_DIR, filename)
                })

    return {"templates": templates}


@router.post("/generate/document")
async def generate_document(request: DocumentRequest):
    """生成法律文书"""
    case_data_list = []
    for case_id in request.case_ids:
        case_data = database_service.load_case(case_id)
        if case_data:
            case_data_list.append(case_data)

    if not case_data_list:
        raise HTTPException(status_code=400, detail="没有选中的案件")

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    if request.doc_type == "起诉状":
        # 生成两份起诉状（两种案由）
        files = []
        for case_data in case_data_list:
            for case_reason in ["清算责任纠纷", "损害公司债权人责任纠纷"]:
                case_data.case_info.case_reason = case_reason
                output_path = os.path.join(OUTPUT_DIR, f"民事起诉状_{case_data.id}_{case_reason[:4]}_{timestamp}.docx")

                # 获取模板变量
                vars = case_data.to_template_vars()

                # 使用模板生成
                template_path = os.path.join(TEMPLATE_DIR, "民事起诉状模板.docx")
                if os.path.exists(template_path):
                    docx_service.generate_from_template(template_path, vars, output_path)
                    files.append(output_path)

        # 打包返回
        if len(files) > 1:
            zip_path = os.path.join(OUTPUT_DIR, f"起诉状_{timestamp}.zip")
            with zipfile.ZipFile(zip_path, 'w') as zipf:
                for f in files:
                    zipf.write(f, os.path.basename(f))
            return FileResponse(zip_path, media_type="application/zip", filename=f"起诉状_{timestamp}.zip")
        elif files:
            return FileResponse(files[0], media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document", filename=os.path.basename(files[0]))

    elif request.doc_type == "证据目录":
        output_path = os.path.join(OUTPUT_DIR, f"证据目录_{timestamp}.docx")
        docx_service.generate_evidence_list(case_data_list[0], output_path)
        return FileResponse(output_path, media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document", filename=os.path.basename(output_path))

    elif request.doc_type == "申请书":
        output_path = os.path.join(OUTPUT_DIR, f"保全申请书_{timestamp}.docx")
        # 使用模板生成
        template_path = os.path.join(TEMPLATE_DIR, "保全申请书模板.docx")
        if os.path.exists(template_path):
            vars = case_data_list[0].to_template_vars()
            docx_service.generate_from_template(template_path, vars, output_path)
        else:
            content = await legal_agent.generate_application(case_data_list[0].model_dump())
            docx_service.generate_application(case_data_list[0], content, output_path)
        return FileResponse(output_path, media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document", filename=os.path.basename(output_path))

    raise HTTPException(status_code=400, detail="不支持的文书类型")


@router.post("/generate/selected")
async def generate_selected(request: SelectedDocumentsRequest):
    """根据选择的文书类型生成文书"""
    case_data_list = []
    for case_id in request.case_ids:
        case_data = database_service.load_case(case_id)
        if case_data:
            case_data_list.append(case_data)

    if not case_data_list:
        raise HTTPException(status_code=400, detail="没有选中的案件")

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    files = []

    # 文书类型映射 - 支持所有15个模板
    doc_type_map = {
        # 民事起诉状（两种案由）
        "起诉状-清算责任纠纷": ("1-民事起诉状模板.docx", "起诉状_清算责任纠纷", True),
        "起诉状-损害公司债权人责任纠纷": ("1-民事起诉状模板.docx", "起诉状_损害公司债权人责任纠纷", True),
        # 其他文书
        "证据目录": ("2-证据目录模板.docx", "证据目录", False),
        "保函": ("3-保函模板.docx", "保函", False),
        "保全申请书": ("4-保全申请书模板.docx", "保全申请书", False),
        "法律文书送达地址确认": ("5-法律文书送达地址确认模板.docx", "法律文书送达地址确认书", False),
        "诉讼授权委托书": ("6-诉讼授权委托书模板.docx", "诉讼授权委托书", False),
        "诉讼文书送达地址确认": ("7-诉讼文书送达地址确认模板.docx", "诉讼文书送达地址确认书", False),
        "公函": ("8-公函模板.docx", "公函", False),
        "诉讼费退费账号": ("9-诉讼费退费账号模板.docx", "诉讼费退费账号确认书", False),
        "网络查控申请书": ("10-网络查控申请书模板.docx", "网络查控申请书", False),
        "执行款收款账户": ("11-执行款收款账户模板.docx", "执行款收款账户确认书", False),
        "执行授权委托书": ("12-执行授权委托书模板.docx", "执行授权委托书", False),
        "执行申请书": ("13-执行申请书模板.docx", "执行申请书", False),
        "律师函": ("14-律师函模板.docx", "律师函", False),
    }

    for case_data in case_data_list:
        for doc_type in request.doc_types:
            if doc_type not in doc_type_map:
                continue

            template_file, output_name, is_complaint = doc_type_map[doc_type]

            # 设置案由
            if is_complaint and "清算责任" in doc_type:
                case_data.case_info.case_reason = "清算责任纠纷"
            elif is_complaint and "损害公司债权人" in doc_type:
                case_data.case_info.case_reason = "损害公司债权人责任纠纷"

            # 生成文书
            template_path = os.path.join(TEMPLATE_DIR, template_file)
            if os.path.exists(template_path):
                # 文件名格式：文书名-公司名.docx
                company_name = case_data.company_info.target_company or case_data.id
                output_filename = f"{output_name}-{company_name}.docx"
                output_path = os.path.join(OUTPUT_DIR, output_filename)
                docx_service.generate_from_template(template_path, case_data.to_template_vars(), output_path)
                files.append(output_path)

    if not files:
        raise HTTPException(status_code=400, detail="没有生成任何文书")

    # 打包为 ZIP
    zip_path = os.path.join(OUTPUT_DIR, f"法律文书_{timestamp}.zip")
    with zipfile.ZipFile(zip_path, 'w') as zipf:
        for f in files:
            zipf.write(f, os.path.basename(f))

    return FileResponse(
        zip_path,
        media_type="application/zip",
        filename=os.path.basename(zip_path)
    )


@router.post("/generate/all")
async def generate_all(case_ids: List[str] = Form(...)):
    """批量生成所有文书"""
    case_data_list = []
    for case_id in case_ids:
        case_data = database_service.load_case(case_id)
        if case_data:
            case_data_list.append(case_data)

    if not case_data_list:
        raise HTTPException(status_code=400, detail="没有选中的案件")

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    files = []

    for case_data in case_data_list:
        company_name = case_data.company_info.target_company or case_data.id

        # 生成起诉状（两种案由）
        for case_reason in ["清算责任纠纷", "损害公司债权人责任纠纷"]:
            case_data.case_info.case_reason = case_reason
            vars = case_data.to_template_vars()

            # 起诉状
            template_path = os.path.join(TEMPLATE_DIR, "1-民事起诉状模板.docx")
            if os.path.exists(template_path):
                output_path = os.path.join(OUTPUT_DIR, f"民事起诉状-{case_reason[:4]}-{company_name}.docx")
                docx_service.generate_from_template(template_path, vars, output_path)
                files.append(output_path)

        # 证据目录
        template_path = os.path.join(TEMPLATE_DIR, "2-证据目录模板.docx")
        if os.path.exists(template_path):
            output_path = os.path.join(OUTPUT_DIR, f"证据目录-{company_name}.docx")
            docx_service.generate_from_template(template_path, case_data.to_template_vars(), output_path)
            files.append(output_path)

        # 保函
        template_path = os.path.join(TEMPLATE_DIR, "3-保函模板.docx")
        if os.path.exists(template_path):
            output_path = os.path.join(OUTPUT_DIR, f"保函-{company_name}.docx")
            docx_service.generate_from_template(template_path, case_data.to_template_vars(), output_path)
            files.append(output_path)

        # 保全申请书
        template_path = os.path.join(TEMPLATE_DIR, "4-保全申请书模板.docx")
        if os.path.exists(template_path):
            output_path = os.path.join(OUTPUT_DIR, f"保全申请书-{company_name}.docx")
            docx_service.generate_from_template(template_path, case_data.to_template_vars(), output_path)
            files.append(output_path)

        # 法律文书送达地址确认书
        template_path = os.path.join(TEMPLATE_DIR, "5-法律文书送达地址确认模板.docx")
        if os.path.exists(template_path):
            output_path = os.path.join(OUTPUT_DIR, f"法律文书送达地址确认书-{company_name}.docx")
            docx_service.generate_from_template(template_path, case_data.to_template_vars(), output_path)
            files.append(output_path)

        # 诉讼授权委托书
        template_path = os.path.join(TEMPLATE_DIR, "6-诉讼授权委托书模板.docx")
        if os.path.exists(template_path):
            output_path = os.path.join(OUTPUT_DIR, f"诉讼授权委托书-{company_name}.docx")
            docx_service.generate_from_template(template_path, case_data.to_template_vars(), output_path)
            files.append(output_path)

        # 诉讼文书送达地址确认书
        template_path = os.path.join(TEMPLATE_DIR, "7-诉讼文书送达地址确认模板.docx")
        if os.path.exists(template_path):
            output_path = os.path.join(OUTPUT_DIR, f"诉讼文书送达地址确认书-{company_name}.docx")
            docx_service.generate_from_template(template_path, case_data.to_template_vars(), output_path)
            files.append(output_path)

        # 公函
        template_path = os.path.join(TEMPLATE_DIR, "8-公函模板.docx")
        if os.path.exists(template_path):
            output_path = os.path.join(OUTPUT_DIR, f"公函-{company_name}.docx")
            docx_service.generate_from_template(template_path, case_data.to_template_vars(), output_path)
            files.append(output_path)

        # 诉讼费退费账号确认书
        template_path = os.path.join(TEMPLATE_DIR, "9-诉讼费退费账号模板.docx")
        if os.path.exists(template_path):
            output_path = os.path.join(OUTPUT_DIR, f"诉讼费退费账号确认书-{company_name}.docx")
            docx_service.generate_from_template(template_path, case_data.to_template_vars(), output_path)
            files.append(output_path)

        # 网络查控申请书
        template_path = os.path.join(TEMPLATE_DIR, "10-网络查控申请书模板.docx")
        if os.path.exists(template_path):
            output_path = os.path.join(OUTPUT_DIR, f"网络查控申请书-{company_name}.docx")
            docx_service.generate_from_template(template_path, case_data.to_template_vars(), output_path)
            files.append(output_path)

        # 执行款收款账户确认书
        template_path = os.path.join(TEMPLATE_DIR, "11-执行款收款账户模板.docx")
        if os.path.exists(template_path):
            output_path = os.path.join(OUTPUT_DIR, f"执行款收款账户确认书-{company_name}.docx")
            docx_service.generate_from_template(template_path, case_data.to_template_vars(), output_path)
            files.append(output_path)

        # 执行授权委托书
        template_path = os.path.join(TEMPLATE_DIR, "12-执行授权委托书模板.docx")
        if os.path.exists(template_path):
            output_path = os.path.join(OUTPUT_DIR, f"执行授权委托书-{company_name}.docx")
            docx_service.generate_from_template(template_path, case_data.to_template_vars(), output_path)
            files.append(output_path)

        # 执行申请书
        template_path = os.path.join(TEMPLATE_DIR, "13-执行申请书模板.docx")
        if os.path.exists(template_path):
            output_path = os.path.join(OUTPUT_DIR, f"执行申请书-{company_name}.docx")
            docx_service.generate_from_template(template_path, case_data.to_template_vars(), output_path)
            files.append(output_path)

        # 律师函
        template_path = os.path.join(TEMPLATE_DIR, "14-律师函模板.docx")
        if os.path.exists(template_path):
            output_path = os.path.join(OUTPUT_DIR, f"律师函-{company_name}.docx")
            docx_service.generate_from_template(template_path, case_data.to_template_vars(), output_path)
            files.append(output_path)

    # 打包为 ZIP
    zip_path = os.path.join(OUTPUT_DIR, f"法律文书_{timestamp}.zip")
    with zipfile.ZipFile(zip_path, 'w') as zipf:
        for f in files:
            zipf.write(f, os.path.basename(f))

    return FileResponse(
        zip_path,
        media_type="application/zip",
        filename=os.path.basename(zip_path)
    )


# ==================== AI助手接口 ====================

class AIChatRequest(BaseModel):
    """AI对话请求"""
    message: str
    history: List[dict] = []
    context: Optional[dict] = None
    case_id: Optional[str] = None
    stream: bool = True


class AIFileAnalyzeRequest(BaseModel):
    """文件分析请求"""
    file_id: str
    question: Optional[str] = None


class AIFormSuggestionRequest(BaseModel):
    """表单建议请求"""
    case_id: str
    field_name: str
    field_label: Optional[str] = ""
    current_value: Optional[str] = None


@router.post("/ai/chat")
async def ai_chat(request: AIChatRequest):
    """AI对话接口 - 支持流式响应"""

    if request.stream:
        # 流式响应
        return StreamingResponse(
            stream_chat_response(
                message=request.message,
                history=request.history,
                context=request.context
            ),
            media_type="text/event-stream"
        )
    else:
        # 非流式响应
        try:
            response = await chat_response(
                message=request.message,
                history=request.history,
                context=request.context
            )
            return {
                "success": True,
                "message": response
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }


@router.post("/ai/analyze-file")
async def ai_analyze_file(request: AIFileAnalyzeRequest):
    """AI分析文件内容"""

    try:
        # 从数据库获取案件文件内容
        # 这里简化处理，实际应该根据file_id读取文件
        file_content = "文件内容..."  # TODO: 实际读取文件

        analysis = await analyze_file_content(
            file_content=file_content,
            question=request.question
        )

        return {
            "success": True,
            "analysis": analysis
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }


@router.post("/ai/form-suggestions")
async def ai_form_suggestions(request: AIFormSuggestionRequest):
    """AI表单填充建议"""

    try:
        # 获取案件数据
        case_data = database_service.load_case(request.case_id)
        if not case_data:
            return {
                "success": False,
                "error": "案件不存在"
            }

        suggestion = await get_form_field_suggestion(
            field_name=request.field_name,
            field_label=request.field_label or request.field_name,
            case_data=case_data.model_dump(),
            current_value=request.current_value
        )

        return {
            "success": True,
            "suggestion": suggestion
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }


@router.post("/ai/check-form/{case_id}")
async def ai_check_form(case_id: str):
    """AI检查表单完整性"""

    try:
        case_data = database_service.load_case(case_id)
        if not case_data:
            return {
                "success": False,
                "error": "案件不存在"
            }

        result = await check_form_completeness(case_data.model_dump())

        return {
            "success": True,
            "check_result": result
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }


# ==================== 批量生成文书接口 ====================


def _parse_batch_xlsx(file) -> list:
    """解析批量上传的 XLSX 文件，返回数据行列表（字典格式）"""
    from openpyxl import load_workbook

    wb = load_workbook(file, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(min_row=1, values_only=True)
    all_rows = [[str(cell) if cell is not None else '' for cell in row] for row in rows_iter]
    wb.close()

    if len(all_rows) < 2:
        raise HTTPException(status_code=400, detail="Excel中没有数据行")

    columns = all_rows[0]
    data_rows = all_rows[1:]
    data_rows = [row for row in data_rows if any(cell.strip() for cell in row)]

    rows = []
    for row in data_rows:
        row_dict = {}
        for i, col in enumerate(columns):
            row_dict[col] = row[i] if i < len(row) else ''
        rows.append(row_dict)
    return rows


async def _do_preview_batch_xlsx(file: UploadFile, doc_type: str):
    """预览批量XLSX的核心逻辑"""
    if not file.filename.lower().endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="仅支持.xlsx格式文件")

    if doc_type not in doc_type_map:
        raise HTTPException(status_code=400, detail=f"不支持的文书类型: {doc_type}")

    temp_path = os.path.join(UPLOAD_DIR, f"temp_batch_{uuid.uuid4().hex[:8]}.xlsx")
    with open(temp_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    rows = _parse_batch_xlsx(temp_path)

    template_file, _, _ = doc_type_map[doc_type]
    template_path = os.path.join(TEMPLATE_DIR, template_file)
    template_variables = []
    if os.path.exists(template_path):
        template_variables = docx_service.extract_template_variables(template_path)

    os.remove(temp_path)

    columns = list(rows[0].keys()) if rows else []

    return {
        "success": True,
        "columns": columns,
        "rows": rows,
        "row_count": len(rows),
        "template_variables": template_variables
    }


async def _do_generate_batch(file: UploadFile, doc_type: str):
    """批量生成文书的核心逻辑"""
    if not file.filename.lower().endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="仅支持.xlsx格式文件")

    if doc_type not in doc_type_map:
        raise HTTPException(status_code=400, detail=f"不支持的文书类型: {doc_type}")

    temp_path = os.path.join(UPLOAD_DIR, f"temp_batch_gen_{uuid.uuid4().hex[:8]}.xlsx")
    with open(temp_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    rows = _parse_batch_xlsx(temp_path)

    template_file, output_name, _ = doc_type_map[doc_type]
    template_path = os.path.join(TEMPLATE_DIR, template_file)
    if not os.path.exists(template_path):
        raise HTTPException(status_code=400, detail=f"模板不存在: {template_file}")

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = os.path.join(OUTPUT_DIR, f"{output_name}_批量生成_{timestamp}.docx")
    docx_service.generate_batch_to_single_docx(template_path, rows, output_path)

    os.remove(temp_path)

    return FileResponse(
        output_path,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        filename=f"{output_name}_批量生成_{timestamp}.docx"
    )


@router.post("/batch/preview-xlsx")
async def preview_batch_xlsx(
    file: UploadFile = File(...),
    doc_type: str = Form(...)
):
    """上传批量数据XLSX，返回表格预览数据和模板变量名

    doc_type: 文书类型（与 doc_type_map 的 key 对应），如 '律师函'、'证据目录' 等
    """
    try:
        return await _do_preview_batch_xlsx(file, doc_type)
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"读取Excel失败: {str(e)}")


@router.post("/batch/generate")
async def generate_batch(
    file: UploadFile = File(...),
    doc_type: str = Form(...)
):
    """上传批量数据XLSX，一键批量生成文书（每份单独一页，合并在一个Word文件中）

    doc_type: 文书类型（与 doc_type_map 的 key 对应），如 '律师函'、'证据目录' 等
    """
    try:
        return await _do_generate_batch(file, doc_type)
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"批量生成文书失败: {str(e)}")


# 旧端点兼容重定向（内部直调，不走 FastAPI 参数解析）
@router.post("/batch/preview-lawyer-letter-xlsx")
async def preview_lawyer_letter_xlsx_legacy(file: UploadFile = File(...)):
    """(兼容旧版) 预览律师函批量数据"""
    return await _do_preview_batch_xlsx(file, "律师函")

@router.post("/batch/generate-lawyer-letters")
async def generate_lawyer_letters_legacy(file: UploadFile = File(...)):
    """(兼容旧版) 批量生成律师函"""
    return await _do_generate_batch(file, "律师函")


@router.get("/company/idcard-image")
async def get_idcard_image(folder: str, name: str):
    """获取身份证图片"""

    # 在文件夹中查找身份证图片
    for filename in os.listdir(folder):
        if name in filename and "身份证" in filename:
            file_path = os.path.join(folder, filename)
            # 返回图片文件
            return FileResponse(
                file_path,
                media_type="image/jpeg"
            )

    raise HTTPException(status_code=404, detail="未找到身份证图片")
