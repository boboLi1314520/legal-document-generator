"""
Excel 生成服务 - 重构版
支持生成 database_element.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from ..models.case import CaseData, CompanyInfo


class ExcelService:
    """Excel 文件生成服务"""

    def generate(self, cases: list, output_path: str):
        """生成案件要素 Excel（旧版兼容）"""
        wb = Workbook()
        ws = wb.active
        ws.title = "案件要素"

        # 表头样式
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=12, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 表头
        headers = ["序号", "文件名", "原告", "被告", "案件类型", "标的额", "案件事实", "证据", "受理法院", "状态", "创建时间"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # 数据行
        for row, case in enumerate(cases, 2):
            ws.cell(row=row, column=1, value=row - 1).border = border
            ws.cell(row=row, column=2, value=case.get("filename", "")).border = border
            ws.cell(row=row, column=3, value=case.get("plaintiff", "")).border = border
            ws.cell(row=row, column=4, value=case.get("defendant", "")).border = border
            ws.cell(row=row, column=5, value=case.get("case_type", "")).border = border
            ws.cell(row=row, column=6, value=case.get("amount", "")).border = border
            ws.cell(row=row, column=7, value=case.get("facts", "")).border = border
            ws.cell(row=row, column=8, value="\n".join(case.get("evidence", []))).border = border
            ws.cell(row=row, column=9, value=case.get("court", "")).border = border
            ws.cell(row=row, column=10, value=case.get("status", "")).border = border
            ws.cell(row=row, column=11, value=case.get("created_at", "")).border = border

            # 设置自动换行
            for col in range(1, 12):
                ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='center')

        # 调整列宽
        column_widths = [6, 20, 15, 15, 12, 12, 40, 30, 15, 10, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        wb.save(output_path)
        return output_path

    def _write_section_header(self, ws, row, title, border):
        """板块标题行：合并单元格、灰底、左侧粗体"""
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border

    def _write_headers(self, ws, row, headers, border):
        """蓝底白字表头行"""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    def _write_data_row(self, ws, row, values, border):
        """普通数据行"""
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value if value is not None else "")
            cell.border = border
            cell.alignment = Alignment(vertical='center')

    def generate_database_element(self, case_data: CaseData, output_path: str) -> str:
        """生成 database_element.xlsx（7板块分区布局）"""
        wb = Workbook()
        ws = wb.active
        ws.title = "要素数据库"

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        ci = case_data.company_info
        di = case_data.debt_info
        csi = case_data.case_info
        row = 1

        # ========== 板块1：公司信息 ==========
        self._write_section_header(ws, row, "公司信息", border)
        row += 1
        self._write_headers(ws, row, [
            "企业名称", "公司简称", "法定代表人", "注册资本", "成立日期",
            "住所", "登记机关", "核准日期", "认缴日期", "出资状态"
        ], border)
        row += 1
        self._write_data_row(ws, row, [
            ci.target_company, ci.company_abbr, ci.legal_representative,
            CompanyInfo.format_capital(ci.company_capital),
            ci.company_establish, ci.company_addr, ci.company_reg,
            ci.company_cancel_apply, ci.subscribe_date, ci.capital_status
        ], border)
        row += 2  # 跳过空行

        # ========== 板块2：被告信息 ==========
        self._write_section_header(ws, row, "被告信息", border)
        row += 1
        self._write_headers(ws, row, [
            "姓名", "性别", "民族", "身份证号", "住址", "电话", "持股比例"
        ], border)
        row += 1
        if case_data.defendants:
            for defendant in case_data.defendants:
                self._write_data_row(ws, row, [
                    defendant.def_name, defendant.def_gender, defendant.def_nation,
                    defendant.def_id, defendant.def_addr, defendant.def_tel, defendant.def_share
                ], border)
                row += 1
        else:
            self._write_data_row(ws, row, [""] * 7, border)
            row += 1
        row += 1  # 跳过空行

        # ========== 板块3：债务信息 ==========
        self._write_section_header(ws, row, "债务信息", border)
        row += 1
        self._write_headers(ws, row, [
            "贷款本金合计", "欠付本金", "利息", "罚息",
            "截止日期", "保全金额", "取整保全金额"
        ], border)
        row += 1
        self._write_data_row(ws, row, [
            di.loan_total, di.principal, di.interest, di.penalty_cutoff,
            di.cutoff_date, di.guarantee_amount, di.guarantee_amount_rounded
        ], border)
        row += 2

        # ========== 板块4：案件信息 ==========
        self._write_section_header(ws, row, "案件信息", border)
        row += 1
        self._write_headers(ws, row, [
            "案由", "受理法院", "案号", "委托律师", "判决文书",
            "放款流水页数", "还款流水页数", "金额计算表页数"
        ], border)
        row += 1
        self._write_data_row(ws, row, [
            csi.case_reason, csi.court_name, csi.case_number, case_data.lawer,
            csi.judgment_document, csi.page_number1, csi.page_number2, csi.page_number3
        ], border)
        row += 2

        # ========== 板块5：额度合同 ==========
        self._write_section_header(ws, row, "额度合同", border)
        row += 1
        self._write_headers(ws, row, ["合同编号", "签订日期"], border)
        row += 1
        quota_contracts = case_data.loan_contracts.quota_contracts
        if quota_contracts:
            for qc in quota_contracts:
                self._write_data_row(ws, row, [qc.contract_no, qc.contract_date], border)
                row += 1
        else:
            self._write_data_row(ws, row, ["", ""], border)
            row += 1
        row += 1

        # ========== 板块6：借款合同明细 ==========
        self._write_section_header(ws, row, "借款合同明细", border)
        row += 1
        self._write_headers(ws, row, [
            "借据号", "合同编号", "签订日期", "本金", "欠款总本金",
            "利率", "罚息利率", "规范罚息利率"
        ], border)
        row += 1
        loan_contracts = case_data.loan_contracts.loan_contracts
        if loan_contracts:
            for loan in loan_contracts:
                self._write_data_row(ws, row, [
                    loan.jieju_no, loan.contract_no, loan.contract_date,
                    loan.principal, loan.total_principal,
                    loan.rate, loan.penalty_rate, loan.standard_penalty_rate
                ], border)
                row += 1
        else:
            self._write_data_row(ws, row, [""] * 8, border)
            row += 1
        row += 1

        # ========== 板块7：借款合同汇总 ==========
        self._write_section_header(ws, row, "借款合同汇总", border)
        row += 1
        self._write_headers(ws, row, [
            "序号", "欠款总本金合计", "规范罚息利率", "合同数"
        ], border)
        row += 1
        template_vars = case_data.to_template_vars()
        loan_summary = template_vars.get("loan_summary", [])
        if loan_summary:
            for idx, item in enumerate(loan_summary, 1):
                self._write_data_row(ws, row, [
                    idx, item.get("principal", ""), item.get("standard_penalty_rate", ""), item.get("count", "")
                ], border)
                row += 1
        else:
            self._write_data_row(ws, row, ["", "", "", ""], border)
            row += 1

        # 设置列宽
        column_widths = [22, 16, 12, 14, 16, 28, 22, 16, 14, 14, 14]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        wb.save(output_path)
        return output_path

    def generate_batch_excel(self, cases: list, output_path: str) -> str:
        """批量生成多个案件的要素Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "要素数据库"

        # 样式定义
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 表头
        headers = [
            "案件ID", "公司名", "法人姓名", "被告一", "被告二",
            "欠付本金", "利息", "罚息", "保全金额",
            "额度合同数", "借款合同数", "贷款本金合计",
            "创建时间", "更新时间"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # 数据行
        for row_idx, case in enumerate(cases, 2):
            ws.cell(row=row_idx, column=1, value=case.id).border = border
            ws.cell(row=row_idx, column=2, value=case.company_info.target_company).border = border
            ws.cell(row=row_idx, column=3, value=case.company_info.legal_representative).border = border

            def1 = case.get_defendant_by_index(1)
            def2 = case.get_defendant_by_index(2)
            ws.cell(row=row_idx, column=4, value=def1.def_name).border = border
            ws.cell(row=row_idx, column=5, value=def2.def_name).border = border

            ws.cell(row=row_idx, column=6, value=case.debt_info.principal).border = border
            ws.cell(row=row_idx, column=7, value=case.debt_info.interest).border = border
            ws.cell(row=row_idx, column=8, value=case.debt_info.penalty_cutoff).border = border
            ws.cell(row=row_idx, column=9, value=case.debt_info.guarantee_amount).border = border

            ws.cell(row=row_idx, column=10, value=case.loan_contracts.quota_count).border = border
            ws.cell(row=row_idx, column=11, value=case.loan_contracts.loan_count).border = border
            ws.cell(row=row_idx, column=12, value=case.debt_info.loan_total).border = border

            ws.cell(row=row_idx, column=13, value=case.created_at).border = border
            ws.cell(row=row_idx, column=14, value=case.updated_at).border = border

        # 设置列宽
        column_widths = [10, 25, 10, 10, 10, 12, 12, 12, 12, 10, 10, 12, 20, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        wb.save(output_path)
        return output_path
