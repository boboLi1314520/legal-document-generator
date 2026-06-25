# -*- coding: utf-8 -*-
"""导出 v1.0 变量手册到 xlsx"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()

header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def write_sheet(ws, title, headers, rows, col_widths=None):
    ws.title = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    for r, row in enumerate(rows, 3):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
    if col_widths:
        for c, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w


# Sheet 1: 被告信息
ws1 = wb.active
write_sheet(ws1, "被告信息变量", ["变量名", "数据来源", "含义"], [
    ["def1_name", "defendants[0].def_name", "第1个被告的姓名"],
    ["def1_gender", "defendants[0].def_gender", "第1个被告的性别"],
    ["def1_nation", "defendants[0].def_nation", "第1个被告的民族"],
    ["def1_id", "defendants[0].def_id", "第1个被告的身份证号码"],
    ["def1_addr", "defendants[0].def_addr", "第1个被告的住址"],
    ["def1_tel", "defendants[0].def_tel", "第1个被告的电话"],
    ["def1_share", "defendants[0].def_share", "第1个被告的持股比例"],
    ["def2_name", "defendants[1].def_name", "第2个被告的姓名（向后兼容）"],
    ["def2_gender", "defendants[1].def_gender", "第2个被告的性别"],
    ["def2_nation", "defendants[1].def_nation", "第2个被告的民族"],
    ["def2_id", "defendants[1].def_id", "第2个被告的身份证号码"],
    ["def2_addr", "defendants[1].def_addr", "第2个被告的住址"],
    ["def2_tel", "defendants[1].def_tel", "第2个被告的电话"],
    ["def2_share", "defendants[1].def_share", "第2个被告的持股比例"],
    ["defendants_text", "_build_defendant_lines('被告', 'full')", "【动态】所有被告完整信息文本块（起诉状用）。格式：被告：姓名，性别，民族族，身份证号码：xxx，住址：xxx，电话：xxx。"],
    ["respondents_text", "_build_defendant_lines('被申请人', 'full')", "【动态】所有被申请人文本块（保全申请书用），前缀为'被申请人'"],
    ["guarantee_defendants_text", "_build_defendant_lines(format_type='guarantee')", "【动态】保函用被告文本块，仅含姓名+身份证号"],
    ["def", "get_defendant_names()", "所有被告姓名用顿号拼接，如'张三、李四'"],
], col_widths=[30, 45, 80])

# Sheet 2: 公司信息
ws2 = wb.create_sheet()
write_sheet(ws2, "公司信息变量", ["变量名", "数据来源", "含义"], [
    ["target_company", "company_info.target_company", "公司全称"],
    ["company_abbr", "company_info.company_abbr", "公司简称"],
    ["company_capital", "CompanyInfo.format_capital()", "注册资本（格式化，如'100万元'）"],
    ["company_establish", "company_info.company_establish", "成立日期"],
    ["company_cancel_apply", "company_info.company_cancel_apply", "核准日期"],
    ["company_cancel_date", "company_info.company_cancel_date", "注销日期"],
    ["company_cancel_approve", "company_info.company_cancel_apply（别名）", "核准日期（模板兼容别名）"],
    ["company_addr", "company_info.company_addr", "住所"],
    ["company_reg", "company_info.company_reg", "登记机关"],
    ["Legal_representative", "company_info.legal_representative", "法定代表人"],
    ["capital_status", "company_info.capital_status", "股东出资状态（未实缴/已实缴）"],
    ["subscribe_date", "company_info.subscribe_date", "股东认缴出资日期"],
], col_widths=[28, 42, 70])

# Sheet 3: 债务信息
ws3 = wb.create_sheet()
write_sheet(ws3, "债务信息变量", ["变量名", "数据来源", "含义"], [
    ["loan_total", "debt_info.loan_total", "贷款本金合计"],
    ["principal", "debt_info.principal", "欠付本金"],
    ["interest", "debt_info.interest", "利息"],
    ["penalty_cutoff", "debt_info.penalty_cutoff", "截止日罚息"],
    ["end_date", "debt_info.end_date", "截止日期"],
    ["start_date", "debt_info.start_date", "起算日期（截止日期+1天）"],
    ["guarantee_amount", "自动计算: principal+interest+penalty_cutoff", "保全金额 = 本金 + 利息 + 罚息"],
    ["guarantee_amount_rounded", "calculate_guarantee_amount_rounded()", "取整保全金额：向下取千位，千位<5截断到万位"],
    ["round-off_guarantee_amount", "同上（别名）", "取整保全金额（模板兼容名）"],
], col_widths=[30, 48, 75])

# Sheet 4: 合同信息
ws4 = wb.create_sheet()
write_sheet(ws4, "合同信息变量", ["变量名", "数据来源", "含义"], [
    ["loan_quota_num", "loan_contracts.quota_count", "额度合同总份数"],
    ["loan_contract_num", "loan_contracts.loan_count", "借款合同总份数"],
    ["quota_dates_text", "所有额度合同日期拼接", "如'2020年7月26日、2021年3月15日'"],
    ["loan_quota_date", "同上（别名）", "额度合同日期文本（模板兼容名）"],
    ["loan_dates_text", "所有借款合同日期拼接", "如'2024年7月8日、2025年1月15日'"],
    ["loan_contract_date", "同上（别名）", "借款合同日期文本（模板兼容名）"],
    ["penalty_text", "loan_summary_list分组生成", "罚息条款文本，如'以10000元为基数，按年利率24.00%的标准；以5000元为基数...'"],
    ["loan_contract_proof", "loan_contracts逐条构建", "每份借款合同描述，格式：'日期签订借款本金为xxx元，年利率为xx%的《借款合同》'"],
    ["loan_flow_proof", "loan_contracts汇总计算", "放款存证描述，格式：'原告于日期向公司账户支付xxx元借款，放款成功。'"],
    ["loan_summary", "loan_summary_list", "借款合同分组汇总列表（对象数组，用于循环渲染）"],
    ["loan_summary_count", "len(loan_summary_list)", "借款合同分组数"],
    ["loan_quota_date1..N", "动态生成", "每份额度合同的日期（N=合同数量）"],
    ["loan_contract_date1..N", "动态生成", "每份借款合同的日期（N=合同数量）"],
    ["principal1..N", "动态生成（按罚息利率分组）", "每组罚息对应的本金"],
    ["principal_rate1..N", "动态生成", "每组罚息对应的利率"],
    ["penalty_cutoff1..N", "动态生成", "每组罚息的截止利率（同principal_rate）"],
], col_widths=[28, 38, 85])

# Sheet 5: 案件及其他
ws5 = wb.create_sheet()
write_sheet(ws5, "案件及其他变量", ["变量名", "数据来源", "含义"], [
    ["case_reason", "case_info.case_reason", "案由（清算责任纠纷 / 损害公司债权人利益责任纠纷）"],
    ["court_name / court", "AI推荐或手动填写", "受理法院（LegalAgent根据被告一住址推荐）"],
    ["lawer", "self.lawer", "律师姓名，默认'赵文'"],
    ["date / 日期", "datetime.now()", "当前日期，格式如'2026年06月24日'"],
    ["case number / 案号", "case_info.case_number", "案号（手动填写）"],
    ["judgment document", "case_info.judgment_document", "判决文书（执行申请书用）"],
    ["judgment_case_number", "execution_info.judgment_case_number", "执行依据判决案号"],
    ["judgment_principal", "execution_info.judgment_principal", "判决本金"],
    ["judgment_interest", "execution_info.judgment_interest", "判决利息"],
    ["court_fee", "execution_info.court_fee", "受理费"],
    ["preservation_fee", "execution_info.preservation_fee", "保全费"],
    ["notice_fee", "execution_info.notice_fee", "公告费"],
    ["page number1", "case_info.page_number1", "证据目录页码1（放款流水），默认'1'"],
    ["page number2", "case_info.page_number2", "证据目录页码2（还款流水），默认'2'"],
    ["page number3", "case_info.page_number3", "证据目录页码3（金额计算逻辑），默认'3'"],
    ["收件人", "lawyer_letter_info.recipient", "律师函收件人"],
    ["金额", "lawyer_letter_info.amount 或 debt_info.principal", "律师函金额"],
], col_widths=[28, 42, 70])

# Sheet 6: 文书模板变量映射
ws6 = wb.create_sheet()
write_sheet(ws6, "文书模板变量映射", ["文书模板", "变量数", "使用的变量"], [
    ["民事起诉状-清算责任纠纷", "27", "defendants_text, def, def1_addr, def1_share, def2_share, target_company, company_abbr, company_addr, company_cancel_approve, company_capital, company_establish, company_reg, capital_status, subscribe_date, principal, interest, penalty_cutoff, penalty_text, loan_total, end_date, start_date, loan_quota_num, loan_quota_date1, loan_quota_date2, loan_contract_num, court, date"],
    ["民事起诉状-损害公司债权人利益责任纠纷", "27", "同上（与清算责任纠纷模板变量完全一致）"],
    ["证据目录", "28", "case_reason, date, def, def1_name, def1_share, def2_name, def2_share, company_abbr, company_cancel_apply, company_cancel_date, subscribe_date, principal, interest, penalty_cutoff, loan_total, end_date, loan_quota_num, loan_quota_date1, loan_quota_date2, loan_contract_num, loan_contract_date, loan_contract_date1, loan_contract_date2, loan_contract_proof, cancel_doc, page number1, page number2, page number3"],
    ["保函", "6", "case_reason, court, date, def, guarantee_defendants_text, round-off_guarantee_amount"],
    ["保全申请书", "6", "case_reason, court, date, def, respondents_text, round-off_guarantee_amount"],
    ["执行申请书", "16", "case_reason, court, date, def, def1_name, def1_gender, def1_nation, def1_id, def1_addr, def1_tel, end_date, judgment_case_number, judgment_principal, judgment_interest, preservation_fee, announcement_fee"],
    ["律师函", "4", "date, judgment_principal, 收件人, 金额"],
    ["公函", "5", "case_reason, court, date, def, lawer"],
    ["网络查控申请书", "4", "case_reason, court, date, def"],
    ["执行款收款账户确认书", "4", "case_reason, court, date, def"],
    ["执行授权委托书", "4", "case_reason, date, def, lawer"],
    ["诉讼费退费账号确认书", "4", "case_reason, court, date, def"],
    ["法律文书送达地址确认书", "4", "case_reason, court, date, def"],
    ["诉讼文书送达地址确认书", "4", "case_reason, court, date, def"],
    ["诉讼授权委托书", "4", "case_reason, date, def, lawer"],
], col_widths=[36, 10, 120])

# Sheet 7: 数据提取流程
ws7 = wb.create_sheet()
write_sheet(ws7, "数据提取流程", ["步骤", "处理对象", "提取内容", "填充目标模型"], [
    ["1", "金额计算逻辑 Excel", "公司名、统一社会信用代码、法定代表人及身份证/电话、债务信息（本金/利息/罚息/截止日期）、借款合同详情（借据号/合同号/本金/利率/罚息率）、保全金额", "company_info + debt_info + LoanContract[] + 被告一(法定代表人)"],
    ["2", "额度合同 PDF", "合同编号、签订日期（从文件名正则提取）", "QuotaContract[]{contract_no, contract_date}"],
    ["3", "借款合同 PDF", "优先用Excel数据（字段更完整），备用方案从PDF文件名提取合同编号和日期", "LoanContract[]{jieju_no, contract_no, contract_date, principal, rate, penalty_rate, ...}"],
    ["4", "公示系统 PDF", "企业名称、注册资本、住所、登记机关、成立日期、核准日期、注销日期、股东姓名+持股比例", "company_info(补充) + defendants(股东信息)"],
    ["5", "身份证 PDF", "性别、民族、身份证号码、住址（OCR识别 + AI Agent解析）", "defendants(匹配股东补充个人信息)"],
    ["6", "判决书 PDF", "判决案号、判决本金、判决利息、受理费、保全费、公告费", "execution_info"],
    ["7", "to_template_vars()", "计算：保全金额(本金+利息+罚息)、取整金额(向下取整规则)、AI推荐法院、当前日期；构建：动态贷款文本/罚息文本/被告文本块", "73+个模板变量字典"],
], col_widths=[8, 24, 75, 70])

desktop = os.path.join(os.environ['USERPROFILE'], 'Desktop')
path = os.path.join(desktop, '法律文书生成系统_v1.0_变量手册.xlsx')
wb.save(path)
print('Saved to: ' + path)
